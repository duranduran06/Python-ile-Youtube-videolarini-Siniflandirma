import os
from openpyxl import Workbook




dirlist = [filename for filename in os.listdir() if os.path.isdir(filename)]
wb = Workbook() 
sheet = wb.active 
c1 = sheet.cell(row = 1, column = 1) 
c1.value = "category"
c2 = sheet.cell(row = 1, column = 2) 
c2.value = "text"

index = 2
for dirName in dirlist:
	os.chdir(dirName)
	textFiles = os.listdir()
	print(dirName)
	for text in textFiles:
		#file= open(text,"r",encoding = "utf-8")
		file= open(text,"r",encoding = "utf-8")
		#data = file.read()
		data = file.readlines()
		file.close()
		cellName = sheet.cell(row = index, column = 1) 
		cellName.value = dirName
		cellText = sheet.cell(row = index, column = 2) 
		cellText.value = data[0]


		index = index + 1
	os.chdir("../")


wb.save(filename="C:/Users/Yasin/Desktop/bi/metin/old texts/yeniv123123.xlsx")
